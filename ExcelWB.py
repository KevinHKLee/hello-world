import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('OriginalBook.xlsx') 
#aaa = wb.sheetnames

##create worksheet
#operations_sheet = wb.create_sheet("Operations",1)
##select worksheet
#products_sheet = workbook["Products"]
##rename worksheet
#products_sheet.title = "New Products"
##remove worksheet
#wb.remove(operations_sheet)


ws = wb.active 
ws['C6'] = 42 

atemp = ws['C5'].value
btemp = ws['C5'].offset(2,0).value

## extract values from cell range
#adf = pd.DataFrame()
#for row in ws.iter_rows(min_row=1,
#                        max_row=8,
#                        min_col=1,
#                        max_col=6,
#                        values_only=True):
#    print(row)
#    adf = adf.append(pd.DataFrame(row).T,ignore_index=True)
#
#adf.fillna(value=pd.np.nan, inplace=True)

## extract rows
#for row in ws.rows:
#    print(row)


wb.save('OutputBook.xlsx') 

print('Output is '+str(atemp)+' '+str(btemp))